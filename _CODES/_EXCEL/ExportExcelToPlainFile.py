#!/usr/bin/python3
# In order to make this file executable in a unix os, execute the code below:
# "chmod 755 GenerateFileWithMetadata.py"

#-------------------------------------------------------------------------------
#  Author        : Cesar R. Urteaga-Reyesvera.
#  Creation date : August 06, 2017.
#-------------------------------------------------------------------------------

#-------------------------------------------------------------------------------
# Creation Date : August 06, 2017.
# Assumptions   : The code assumes that the dimensions of the Excel files are
#                 correct (i.e., there are no empty spaces).
# Description   : Creates a plain file separated by a given character based on
#                 an Excel file.  The program extracts just the corresponding
#                 information of the columns with the specified headers and
#                 could rename the columns.
# Parameters    : Excel_file_name          - File name of the excel file with
#                                            the information.  It could contain
#                                            the absolute path of the file.
#                 worksheet_name           - Name of the worksheet where the
#                                            data is stored.
#                 created_file             - File name of the plain file with
#                                            the data requested.  It could
#                                            specify the path.
#                 targetnames_creatednames - Nested tuples, where each tuple
#                                            must be of the form:
#                                            ("target-column name",
#                                             "created-column name")
#                 lookup_percentage        - A float value between 0 and 1.
#                 character_separator      - The character separator that you
#                                            want to use to delimit each datum.
# Output        : Creates a plain file with the requested information from the
#                 given Excel file delimited by the specified character.
#-------------------------------------------------------------------------------
def ExportExcelToPlainFile(Excel_file_name,
                           worksheet_name,
                           created_file,
                           targetnames_creatednames,
                           lookup_percentage,
                           character_separator):
  # Required libraries:
  from openpyxl import load_workbook
  from difflib  import SequenceMatcher
  import os
  # Reads the workbook that is expected to have the information of the headers
  # above.
  wb = load_workbook(filename  = Excel_file_name,
                     read_only = True)
  ws = wb[worksheet_name]
  # Stores the cells with the headers (i.e., the first row).
  headers = ws[1]
  print("Getting the column numbers.")
  # Gets the column numbers of the target names within the file.
  positions = {}
  for target, created in targetnames_creatednames:
    # If the target header is equal, it stores the column number.
    counter = 0
    for cell in headers:
      if cell.value == target:
        positions[created] = counter
        break
      counter += 1
    # If the header does not match, it will return the column that matches the
    # most the target header.
    if target not in [cell.value for cell in headers]:
      counter = 0
      maximum_value = 0
      for cell in headers:
        distance_ratio = SequenceMatcher(isjunk = None,
                                         a      = cell.value.upper(),
                                         b      = target).ratio()
        if distance_ratio > lookup_percentage and \
           distance_ratio > maximum_value:
          maximum_value = distance_ratio
          positions[created] = counter
        counter += 1
  print("Creating the plain file.")
  # Creates a plain file with the information of the founded columns.
  fout = open(created_file, "w")
  #   Print the headers.
  fout.write(character_separator.join(list(positions.keys())))
  for i in range(2, ws.max_row + 1):
    temporal_list = []
    for j in positions.values():
      temporal_list.append(str(ws.cell(row = i, column = j + 1).value).strip())
    fout.write(character_separator.join(temporal_list))
  fout.close()
  # Prints the absolute file of the created file.
  created_file_path = os.path.abspath(created_file)
  print("The plain file was:", created_file_path)
